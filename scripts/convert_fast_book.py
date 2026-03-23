#!/usr/bin/env python3
"""
Convert FAST Book Part II (Excel) to a JSON reference file for TAS resolution.

The FAST Book (Federal Account Symbols and Titles) is published by the Bureau
of the Fiscal Service at https://tfx.treasury.gov/reference-books/fast-book.
Part II contains all appropriation and fund account symbols and titles.

This script reads the Excel file and produces a clean JSON file suitable for
shipping with the tool as bundled reference data.

Usage:
    python scripts/convert_fast_book.py [--input FILE] [--output FILE]

Default input:  tmp/fast_book_part_ii_iii.xlsx
Default output: data/fas_reference.json
"""

import argparse
import json
import os
import re
import sys
import time

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def normalize_fund_type(raw: str) -> str:
    """Normalize fund type strings to a consistent lowercase slug."""
    lower = raw.strip().lower()
    if "general" in lower:
        return "general"
    elif "revolving" in lower:
        return "revolving"
    elif "special" in lower:
        return "special"
    elif "trust" in lower:
        return "trust"
    elif "deposit" in lower:
        return "deposit"
    elif "management" in lower:
        return "management"
    elif "consolidated" in lower:
        return "consolidated_working"
    else:
        return lower if lower else "unknown"


def parse_fas_code(tas_str: str) -> tuple[str, str, bool] | None:
    """
    Parse a TAS string into (agency_code, main_account, is_no_year).

    TAS strings look like:
        '070 0400'   → annual account
        '070X0400'   → no-year account
        '021 2020'   → annual account
        '070 0400.001' → sub-account (we strip the sub-account)

    Returns (agency_code, main_account, is_no_year) or None if unparseable.
    """
    if not tas_str or not tas_str.strip():
        return None

    cleaned = tas_str.strip()

    # Match patterns: NNN NNNN, NNNXNNNN, NNN NNNN.NNN
    m = re.match(r"(\d{3})\s*([X ]?)\s*(\d{3,4})(?:\.\d+)?", cleaned)
    if not m:
        return None

    agency_code = m.group(1)
    x_flag = m.group(2).strip()
    main_account = m.group(3).zfill(4)  # Pad to 4 digits
    is_no_year = x_flag == "X"

    return (agency_code, main_account, is_no_year)


def parse_part_ii(ws) -> list[dict]:
    """Parse the 'Part II' sheet into a list of account dicts."""
    accounts = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        tas_str = str(row[3] or "").strip()
        if not tas_str:
            continue

        parsed = parse_fas_code(tas_str)
        if parsed is None:
            continue

        agency_code, main_account, is_no_year = parsed
        agency_name = str(row[4] or "").strip()
        title = str(row[5] or "").strip()
        legislation = str(row[6] or "").strip() or None
        fund_type_raw = str(row[7] or "").strip()
        independent = str(row[8] or "").strip() or None
        last_update = str(row[9] or "").strip()

        # Parse last_update to a clean date string
        if last_update and last_update != "None":
            last_update = last_update[:10]  # Just YYYY-MM-DD
        else:
            last_update = None

        fas_code = f"{agency_code}-{main_account}"

        accounts.append({
            "fas_code": fas_code,
            "agency_code": agency_code,
            "main_account": main_account,
            "tas_string": tas_str,
            "agency_name": agency_name,
            "title": title,
            "legislation": legislation,
            "fund_type": normalize_fund_type(fund_type_raw),
            "is_no_year": is_no_year,
            "independent_agency": independent,
            "last_updated": last_update,
        })

    return accounts


def parse_changes(ws) -> list[dict]:
    """Parse the 'Changes' sheet into a list of change records."""
    changes = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        tas_str = str(row[3] or "").strip()
        if not tas_str:
            continue

        parsed = parse_fas_code(tas_str)
        if parsed is None:
            continue

        agency_code, main_account, is_no_year = parsed
        agency_name = str(row[4] or "").strip()
        title = str(row[5] or "").strip()
        legislation = str(row[6] or "").strip() or None
        fund_type_raw = str(row[7] or "").strip()
        last_update = str(row[9] or "").strip()
        action = str(row[10] or "").strip()
        comments = str(row[11] or "").strip() or None

        if last_update and last_update != "None":
            last_update = last_update[:10]
        else:
            last_update = None

        fas_code = f"{agency_code}-{main_account}"

        changes.append({
            "fas_code": fas_code,
            "agency_code": agency_code,
            "main_account": main_account,
            "tas_string": tas_str,
            "agency_name": agency_name,
            "title": title,
            "legislation": legislation,
            "fund_type": normalize_fund_type(fund_type_raw),
            "is_no_year": is_no_year,
            "action": action.lower() if action else None,
            "comments": comments,
            "last_updated": last_update,
        })

    return changes


def collapse_to_fas(accounts: list[dict]) -> list[dict]:
    """
    Collapse TAS-level entries to FAS (Federal Account Symbol) level.

    Multiple TAS entries can share the same FAS code but differ in
    availability type (annual vs no-year). We collapse them into one
    entry per FAS code, keeping the most informative title and noting
    whether no-year variants exist.
    """
    by_fas: dict[str, dict] = {}

    for acct in accounts:
        fas = acct["fas_code"]
        if fas not in by_fas:
            by_fas[fas] = {
                "fas_code": fas,
                "agency_code": acct["agency_code"],
                "main_account": acct["main_account"],
                "agency_name": acct["agency_name"],
                "title": acct["title"],
                "legislation": acct["legislation"],
                "fund_type": acct["fund_type"],
                "has_no_year_variant": acct["is_no_year"],
                "has_annual_variant": not acct["is_no_year"],
                "independent_agency": acct.get("independent_agency"),
                "last_updated": acct["last_updated"],
                "tas_variants": 1,
            }
        else:
            existing = by_fas[fas]
            existing["tas_variants"] += 1
            if acct["is_no_year"]:
                existing["has_no_year_variant"] = True
            else:
                existing["has_annual_variant"] = True
            # Keep the longer/more informative title
            if len(acct["title"]) > len(existing["title"]):
                existing["title"] = acct["title"]
            # Keep the most recent last_updated
            if acct["last_updated"] and (
                not existing["last_updated"]
                or acct["last_updated"] > existing["last_updated"]
            ):
                existing["last_updated"] = acct["last_updated"]
            # Keep legislation if we don't have one
            if acct["legislation"] and not existing["legislation"]:
                existing["legislation"] = acct["legislation"]

    return sorted(by_fas.values(), key=lambda a: a["fas_code"])


def build_discontinued(changes: list[dict], active_fas: set[str]) -> list[dict]:
    """
    Extract discontinued accounts from the Changes sheet that are NOT in
    the active Part II accounts. These are historical accounts that may
    appear in older appropriations bills.
    """
    discontinued = {}

    for change in changes:
        action = (change.get("action") or "").lower()
        if "discontinu" not in action and "deleted" not in action and "removed" not in action:
            continue

        fas = change["fas_code"]
        if fas in active_fas:
            continue  # Still active, not truly discontinued

        if fas not in discontinued:
            discontinued[fas] = {
                "fas_code": fas,
                "agency_code": change["agency_code"],
                "main_account": change["main_account"],
                "agency_name": change["agency_name"],
                "title": change["title"],
                "fund_type": change["fund_type"],
                "action": action,
                "comments": change.get("comments"),
                "last_updated": change["last_updated"],
            }

    return sorted(discontinued.values(), key=lambda a: a["fas_code"])


def main():
    parser = argparse.ArgumentParser(
        description="Convert FAST Book Part II (Excel) to JSON reference file"
    )
    parser.add_argument(
        "--input",
        default="tmp/fast_book_part_ii_iii.xlsx",
        help="Path to FAST Book Excel file",
    )
    parser.add_argument(
        "--output",
        default="data/fas_reference.json",
        help="Path for output JSON file",
    )
    args = parser.parse_args()

    # Resolve paths relative to project root
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    input_path = os.path.join(project_root, args.input)
    output_path = os.path.join(project_root, args.output)

    if not os.path.isfile(input_path):
        print(f"Error: FAST Book file not found: {input_path}")
        print("Download it from: https://tfx.treasury.gov/reference-books/fast-book")
        print('  curl -sL "https://tfx.treasury.gov/media/60111/download?inline" -o tmp/fast_book_part_ii_iii.xlsx')
        sys.exit(1)

    print(f"Reading {input_path}...")
    wb = openpyxl.load_workbook(input_path)

    # Parse Part II (active accounts)
    print("Parsing Part II (appropriation and fund accounts)...")
    raw_accounts = parse_part_ii(wb["Part II"])
    print(f"  Raw TAS entries: {len(raw_accounts)}")

    # Collapse to FAS level
    fas_accounts = collapse_to_fas(raw_accounts)
    print(f"  Collapsed to FAS level: {len(fas_accounts)} unique accounts")

    # Count by fund type
    by_fund = {}
    for a in fas_accounts:
        ft = a["fund_type"]
        by_fund[ft] = by_fund.get(ft, 0) + 1
    for ft, count in sorted(by_fund.items(), key=lambda x: -x[1]):
        print(f"    {ft}: {count}")

    # Parse Changes sheet for discontinued accounts
    print("Parsing Changes sheet (discontinued accounts)...")
    raw_changes = parse_changes(wb["Changes"])
    print(f"  Raw change entries: {len(raw_changes)}")

    active_fas_codes = {a["fas_code"] for a in fas_accounts}
    discontinued = build_discontinued(raw_changes, active_fas_codes)
    general_discontinued = [d for d in discontinued if d["fund_type"] == "general"]
    print(f"  Discontinued accounts not in Part II: {len(discontinued)}")
    print(f"  Discontinued General Fund accounts: {len(general_discontinued)}")

    # Collect unique agency names for reference
    agency_names = {}
    for a in fas_accounts:
        code = a["agency_code"]
        if code not in agency_names:
            agency_names[code] = a["agency_name"]

    # Build output
    output = {
        "schema_version": "1.0",
        "source": "FAST Book Part II — Federal Account Symbols and Titles",
        "source_url": "https://tfx.treasury.gov/reference-books/fast-book",
        "publisher": "Bureau of the Fiscal Service, U.S. Department of the Treasury",
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "description": (
            "Reference data for mapping appropriation provisions to Treasury Account Symbols. "
            "Each entry is a Federal Account Symbol (FAS) — the agency code + main account code "
            "that identifies an appropriation account as established by Congress. FAS codes are "
            "stable through account renames and persist for the life of the account. "
            "Use the 'general' fund_type entries for matching provisions in appropriations bills."
        ),
        "statistics": {
            "total_fas_codes": len(fas_accounts),
            "active_general_fund": sum(
                1 for a in fas_accounts if a["fund_type"] == "general"
            ),
            "discontinued_general_fund": len(general_discontinued),
            "unique_agencies": len(agency_names),
            "by_fund_type": by_fund,
        },
        "agencies": [
            {"code": code, "name": name}
            for code, name in sorted(agency_names.items())
        ],
        "accounts": fas_accounts,
        "discontinued": general_discontinued,
    }

    # Write output
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    file_size = os.path.getsize(output_path)
    print(f"\nWritten {output_path} ({file_size / 1024:.0f} KB)")
    print(f"  {len(fas_accounts)} active FAS codes")
    print(f"  {len(general_discontinued)} discontinued General Fund accounts")
    print(f"  {len(agency_names)} agencies")

    # Summary for verification
    general = [a for a in fas_accounts if a["fund_type"] == "general"]
    print(f"\nGeneral Fund accounts (the ones that appear in appropriations bills):")
    print(f"  Active:       {len(general)}")
    print(f"  Discontinued: {len(general_discontinued)}")
    print(f"  Total known:  {len(general) + len(general_discontinued)}")


if __name__ == "__main__":
    main()
